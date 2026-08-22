from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from project_paths import RESUME_TEMPLATE_PATH, SITE_CONTENT_PATH

ROOT = SITE_CONTENT_PATH.parent.parent


def main() -> None:
    data = json.loads(SITE_CONTENT_PATH.read_text(encoding="utf-8"))
    workbook = load_workbook(RESUME_TEMPLATE_PATH, data_only=False)
    if "_resume_map" not in workbook.sheetnames:
        raise RuntimeError("resume_template.xlsx is missing the hidden _resume_map sheet")

    resume = data["resume"]
    meta = workbook["_resume_map"]

    section_items: dict[tuple[int, int], dict[int, dict]] = {}
    skill_groups: dict[tuple[int, int], dict[int, dict]] = {}
    list_items: dict[tuple[int, int], dict[int, dict]] = {}
    excel_layout: list[dict] = []

    for row in meta.iter_rows(min_row=2, values_only=True):
        sheet_name, page_index, block_index, block_type, field, item_index, group_index, anchor = row
        if not sheet_name or sheet_name not in workbook.sheetnames:
            continue

        value, merged_range = read_anchor_value(workbook[sheet_name], str(anchor))
        page_index = int(page_index)
        excel_layout.append(
            {
                "sheet": sheet_name,
                "page_index": page_index,
                "block_index": "" if block_index is None else block_index,
                "field": field,
                "item_index": "" if item_index is None else item_index,
                "group_index": "" if group_index is None else group_index,
                "anchor": anchor,
                "range": merged_range,
            }
        )

        if field == "name":
            resume["name"] = value
        elif field == "headline":
            resume["headline"] = value
        elif field == "contact":
            resume["contact"] = split_contact(value)
        elif field == "page_title":
            resume["pages"][page_index]["title"] = value
        elif field == "block_heading":
            resume["pages"][page_index]["blocks"][int(block_index)]["heading"] = value
        elif str(block_type) == "skills":
            key = (page_index, int(block_index))
            group = skill_groups.setdefault(key, {}).setdefault(int(group_index), {"label": "", "items": []})
            if field == "skill_label":
                group["label"] = value
            elif field == "skill_items":
                group["items"] = split_items(value)
        elif str(block_type) == "list":
            key = (page_index, int(block_index))
            item = list_items.setdefault(key, {}).setdefault(int(item_index), {"text": ""})
            if field == "list_item":
                item["text"] = value
        else:
            key = (page_index, int(block_index))
            item = section_items.setdefault(key, {}).setdefault(
                int(item_index),
                default_item(resume, page_index, int(block_index), int(item_index)),
            )
            if str(field).startswith("bullet_"):
                bullet_index = int(str(field).split("_", 1)[1])
                while len(item["bullets"]) <= bullet_index:
                    item["bullets"].append("")
                item["bullets"][bullet_index] = value
            elif field in item:
                item[field] = value

    for (page_index, block_index), items in section_items.items():
        block = resume["pages"][page_index]["blocks"][block_index]
        block["items"] = [items[index] for index in sorted(items)]

    for (page_index, block_index), groups in skill_groups.items():
        block = resume["pages"][page_index]["blocks"][block_index]
        block["groups"] = [groups[index] for index in sorted(groups)]

    for (page_index, block_index), items in list_items.items():
        block = resume["pages"][page_index]["blocks"][block_index]
        block["items"] = [items[index] for index in sorted(items)]

    data["resume"] = resume
    data["resume"]["excel_layout"] = excel_layout
    SITE_CONTENT_PATH.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")

    subprocess.run([sys.executable, str(ROOT / "scripts" / "build_site.py")], cwd=ROOT, check=True)
    print(f"Synced {RESUME_TEMPLATE_PATH.relative_to(ROOT)} into {SITE_CONTENT_PATH.relative_to(ROOT)} and rebuilt generated files")


def read_anchor_value(ws, anchor: str) -> tuple[str, str]:
    # The value lives in the anchor cell. The current merged range is intentionally
    # discovered here so users can resize placeholders without editing metadata.
    for merged_range in ws.merged_cells.ranges:
        if anchor in merged_range:
            return clean(ws[merged_range.coord.split(":")[0]].value), merged_range.coord
    return clean(ws[anchor].value), anchor


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def split_contact(value: str) -> list[str]:
    return [part.strip() for part in value.replace("\n", "|").split("|") if part.strip()]


def split_items(value: str) -> list[str]:
    separators = value.replace("\n", ",").split(",")
    return [part.strip() for part in separators if part.strip()]


def default_item(resume: dict, page_index: int, block_index: int, item_index: int) -> dict:
    existing_items = resume["pages"][page_index]["blocks"][block_index].get("items", [])
    existing = existing_items[item_index] if item_index < len(existing_items) else {}
    item = {
        "role": "",
        "organization": "",
        "location": "",
        "dates": "",
        "bullets": [],
    }
    if existing.get("id"):
        item["id"] = existing["id"]
    return item


if __name__ == "__main__":
    main()
