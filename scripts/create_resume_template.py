from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CONTENT_PATH = ROOT / "content" / "site.json"
TEMPLATE_PATH = ROOT / "content" / "resume_template.xlsx"

PAGE_SPECS = [
    {
        "sheet": "Page 1 - Experience",
        "page_index": 0,
        "title": "Work Experience, Skills & Volunteering",
        "blocks": [
            ("section", "Work Experience", 2),
            ("skills", "Skills", 0),
            ("section", "Volunteering & Leadership", 1),
        ],
    },
    {
        "sheet": "Page 2 - Portfolio",
        "page_index": 1,
        "title": "Project & Club Portfolio",
        "blocks": [
            ("section", "Engineering Projects", 2),
            ("section", "Clubs & Teams", 1),
        ],
    },
]


THIN = Side(style="thin", color="D8D3CA")
MEDIUM = Side(style="medium", color="151515")
ACCENT = "E94F2F"
SURFACE = "E9E6DF"
BACKGROUND = "F4F2ED"
INPUT = "FFF8E8"
HELP = "F8FAFC"


def main() -> None:
    data = json.loads(CONTENT_PATH.read_text(encoding="utf-8"))
    resume = data["resume"]

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    meta = wb.create_sheet("_resume_map")
    meta.sheet_state = "hidden"
    meta.append(["sheet", "page_index", "block_index", "block_type", "field", "item_index", "group_index", "anchor"])

    instructions = wb.create_sheet("Instructions")
    build_instructions_sheet(instructions)

    for spec in PAGE_SPECS:
        ws = wb.create_sheet(spec["sheet"])
        build_page_sheet(wb, ws, meta, resume, spec)

    wb.active = 1
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TEMPLATE_PATH)
    print(f"Created {TEMPLATE_PATH.relative_to(ROOT)}")


def build_instructions_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 95

    ws.merge_cells("A1:B1")
    ws["A1"] = "Resume Template Workflow"
    ws["A1"].font = Font(name="Aptos Display", size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="151515")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    rows = [
        ("Edit text", "Replace the bracketed placeholders in the yellow merged cells."),
        ("Resize blocks", "Change merged-cell boundaries in Excel to reserve more or less space for a field. The sync reads the current merged range for each anchor cell."),
        ("Preserve anchors", "Do not delete the top-left anchor cell of a placeholder block. You can change the text and resize the merged area."),
        ("Sync to JSON", "Run: python scripts/sync_resume_from_excel.py"),
        ("Build PDF/site", "Run: python scripts/build_site.py"),
        ("Photos", "Resume sync ignores website photos and project images. The resume never includes photos."),
    ]
    for row, values in enumerate(rows, start=3):
        ws.cell(row=row, column=1, value=values[0])
        ws.cell(row=row, column=2, value=values[1])
        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=HELP)
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=1).font = Font(bold=True, color="151515")
        ws.row_dimensions[row].height = 42


def build_page_sheet(wb: Workbook, ws, meta, resume: dict, spec: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = "portrait"
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    widths = [16, 16, 16, 16, 16, 16, 16, 16]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    page_index = spec["page_index"]
    page = resume["pages"][page_index]

    merge_value(ws, "A1:H1", resume["name"], "title")
    merge_value(ws, "A2:H2", resume["headline"], "headline")
    merge_value(ws, "A3:H3", " | ".join(resume["contact"]), "contact")
    merge_value(ws, "A5:H5", page["title"], "page_title")
    add_meta(meta, ws.title, page_index, "", "", "name", "", "", "A1")
    add_meta(meta, ws.title, page_index, "", "", "headline", "", "", "A2")
    add_meta(meta, ws.title, page_index, "", "", "contact", "", "", "A3")
    add_meta(meta, ws.title, page_index, "", "", "page_title", "", "", "A5")

    style_header(ws, "A1:H1", 18)
    style_subtle(ws, "A2:H3")
    style_page_title(ws, "A5:H5")

    row = 7
    for block_index, block in enumerate(page["blocks"]):
        row = add_section_header(ws, meta, page_index, block_index, block, row)
        if block.get("type") == "skills":
            row = add_skills_block(ws, meta, page_index, block_index, block, row)
        else:
            row = add_items_block(ws, meta, page_index, block_index, block, row)
        row += 1

    apply_print_area(ws, row)


def add_section_header(ws, meta, page_index: int, block_index: int, block: dict, row: int) -> int:
    merge_ref = f"A{row}:H{row}"
    merge_value(ws, merge_ref, block["heading"], "section_heading")
    style_section(ws, merge_ref)
    add_meta(meta, ws.title, page_index, block_index, block.get("type", "section"), "block_heading", "", "", f"A{row}")
    return row + 1


def add_items_block(ws, meta, page_index: int, block_index: int, block: dict, row: int) -> int:
    for item_index, item in enumerate(block.get("items", [])):
        merge_value(ws, f"A{row}:D{row}", item.get("role", ""), "role")
        merge_value(ws, f"E{row}:H{row}", item.get("dates", ""), "dates")
        add_meta(meta, ws.title, page_index, block_index, block.get("type", "section"), "role", item_index, "", f"A{row}")
        add_meta(meta, ws.title, page_index, block_index, block.get("type", "section"), "dates", item_index, "", f"E{row}")
        style_input(ws, f"A{row}:D{row}", bold=True)
        style_input(ws, f"E{row}:H{row}", align="right")
        row += 1

        merge_value(ws, f"A{row}:D{row}", item.get("organization", ""), "organization")
        merge_value(ws, f"E{row}:H{row}", item.get("location", ""), "location")
        add_meta(meta, ws.title, page_index, block_index, block.get("type", "section"), "organization", item_index, "", f"A{row}")
        add_meta(meta, ws.title, page_index, block_index, block.get("type", "section"), "location", item_index, "", f"E{row}")
        style_input(ws, f"A{row}:D{row}")
        style_input(ws, f"E{row}:H{row}", align="right")
        row += 1

        for bullet_index, bullet in enumerate(item.get("bullets", [])):
            merge_value(ws, f"A{row}:H{row + 1}", bullet, "bullet")
            add_meta(meta, ws.title, page_index, block_index, block.get("type", "section"), f"bullet_{bullet_index}", item_index, "", f"A{row}")
            style_input(ws, f"A{row}:H{row + 1}")
            ws.row_dimensions[row].height = 24
            ws.row_dimensions[row + 1].height = 24
            row += 2
        row += 1
    return row


def add_skills_block(ws, meta, page_index: int, block_index: int, block: dict, row: int) -> int:
    for group_index, group in enumerate(block.get("groups", [])):
        merge_value(ws, f"A{row}:B{row}", group.get("label", ""), "skill_label")
        merge_value(ws, f"C{row}:H{row + 1}", ", ".join(group.get("items", [])), "skill_items")
        add_meta(meta, ws.title, page_index, block_index, "skills", "skill_label", "", group_index, f"A{row}")
        add_meta(meta, ws.title, page_index, block_index, "skills", "skill_items", "", group_index, f"C{row}")
        style_input(ws, f"A{row}:B{row}", bold=True)
        style_input(ws, f"C{row}:H{row + 1}")
        ws.row_dimensions[row].height = 24
        ws.row_dimensions[row + 1].height = 24
        row += 2
    return row


def merge_value(ws, range_ref: str, value: str, comment_label: str) -> None:
    ws.merge_cells(range_ref)
    cell = ws[range_ref.split(":")[0]]
    cell.value = value
    cell.comment = Comment(f"{comment_label}: edit text here; resize merged range if this content needs more or less room.", "Codex")
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def style_header(ws, range_ref: str, size: int) -> None:
    style_range(ws, range_ref, fill="151515", font_color="FFFFFF", bold=True, size=size, border=MEDIUM)


def style_page_title(ws, range_ref: str) -> None:
    style_range(ws, range_ref, fill=BACKGROUND, font_color=ACCENT, bold=True, size=12, border=MEDIUM)


def style_section(ws, range_ref: str) -> None:
    style_range(ws, range_ref, fill=SURFACE, font_color="151515", bold=True, size=11, border=MEDIUM)


def style_subtle(ws, range_ref: str) -> None:
    style_range(ws, range_ref, fill=BACKGROUND, font_color="555555", bold=False, size=10, border=THIN)


def style_input(ws, range_ref: str, bold: bool = False, align: str = "left") -> None:
    style_range(ws, range_ref, fill=INPUT, font_color="151515", bold=bold, size=9, border=THIN, align=align)


def style_range(
    ws,
    range_ref: str,
    fill: str,
    font_color: str,
    bold: bool,
    size: int,
    border: Side,
    align: str = "left",
) -> None:
    for row in ws[range_ref]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Aptos", size=size, bold=bold, color=font_color)
            cell.border = Border(left=border, right=border, top=border, bottom=border)
            cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)


def add_meta(meta, sheet: str, page_index, block_index, block_type, field, item_index, group_index, anchor: str) -> None:
    meta.append([sheet, page_index, block_index, block_type, field, item_index, group_index, anchor])


def apply_print_area(ws, last_row: int) -> None:
    last_row = min(max(last_row, 20), 55)
    ws.print_area = f"A1:H{last_row}"
    for row in range(1, last_row + 1):
        if ws.row_dimensions[row].height is None:
            ws.row_dimensions[row].height = 20


if __name__ == "__main__":
    main()
